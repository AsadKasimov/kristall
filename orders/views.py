from django import forms
from datetime import datetime
from openpyxl import Workbook
from django.http import HttpResponse

from django.utils import timezone
from .models import Employee

from .models import Order, Rug, Notification
from .models import OrderStaff
from .forms import OrderForm, AssignEmployeeForm
from django.forms import modelformset_factory
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.core.exceptions import PermissionDenied
from .forms import ClientCreateForm
from .models import AccessToken
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Q
from django.http import JsonResponse
from accounts.models import CustomUser
from django.forms import formset_factory
from .forms import RugForm
from django.utils.timezone import now, localtime
from django.utils.timezone import localdate
from django.db.models import Q
from datetime import datetime, timedelta, date
from calendar import monthrange


def client_dashboard(request, token):
    access_token = get_object_or_404(AccessToken, token=token)
    client = access_token.client
    orders = client.orders.all().order_by('-created_at')
    notifications = client.notifications.order_by('-created_at')

    return render(request, 'orders/dashboard.html', {
        'client': client,
        'orders': orders,
        'notifications': notifications,
    })


def faq_view(request):
    return render(request, 'orders/faq.html')


@login_required
def create_order(request):
    RugFormSet = modelformset_factory(Rug, form=RugForm, extra=2)

    if request.method == 'POST':
        form = OrderForm(request.POST)
        rug_forms = RugFormSet(request.POST, queryset=Rug.objects.none())
        emp_form = AssignEmployeeForm(request.POST)

        if form.is_valid() and rug_forms.is_valid() and emp_form.is_valid():
            order = form.save()
            for rug_form in rug_forms:
                rug = rug_form.save(commit=False)
                rug.order = order
                rug.save()
            OrderStaff.objects.create(order=order, employee=emp_form.cleaned_data['employee'])

            # создаём токен
            AccessToken.objects.get_or_create(client=order.client)

            return redirect('order_success', order_id=order.id)

    else:
        form = OrderForm()
        rug_forms = RugFormSet(queryset=Rug.objects.none())
        emp_form = AssignEmployeeForm()

    return render(request, 'orders/create_order.html', {
        'form': form,
        'rug_forms': rug_forms,
        'emp_form': emp_form,
    })


@login_required
def order_success(request, order_id):
    order = Order.objects.get(id=order_id)
    token = AccessToken.objects.get(client=order.client)
    url = request.build_absolute_uri(f"/d/{token.short_token}/")
    return HttpResponse(f"✅ Заказ создан! Ссылка для клиента: <a href='{url}'>{url}</a>")


@login_required
def operator_order_list(request):
    orders = Order.objects.select_related('client').prefetch_related('rugs').order_by('-created_at')
    return render(request, 'orders/operator_list.html', {'orders': orders})


@login_required
def operator_order_list(request):
    status_filter = request.GET.get('status')
    search = request.GET.get('q')

    orders = Order.objects.select_related('client').prefetch_related('rugs').order_by('-created_at')

    if status_filter:
        orders = orders.filter(status=status_filter)

    if search:
        orders = orders.filter(
            Q(client__username__icontains=search) |
            Q(client__phone__icontains=search)
        )

    statuses = Order.objects.values_list('status', flat=True).distinct()

    return render(request, 'orders/operator_list.html', {
        'orders': orders,
        'statuses': statuses,
        'active_status': status_filter,
        'search': search,
    })


def only_operators(view_func):
    def wrapped(request, *args, **kwargs):
        if request.user.role != 'OPERATOR':
            raise PermissionDenied
        return view_func(request, *args, **kwargs)

    return wrapped


def index(request):
    return render(request, 'index.html')


def token_redirect(request):
    token = request.GET.get('token')
    if token:
        return redirect(f"/dashboard/{token}/")
    return redirect('/')


from django.shortcuts import render


def faq_view(request):
    return render(request, 'orders/faq.html')


@login_required
@only_operators
def operator_dashboard(request):
    return render(request, 'orders/operator_dashboard.html')


@login_required
@only_operators
def create_order_with_client(request, client_id):
    client = get_object_or_404(CustomUser, id=client_id, role='CLIENT')

    if request.method == 'POST':
        form = OrderForm(request.POST)
        if form.is_valid():
            order = form.save(commit=False)
            order.client = client
            order.save()
            return redirect('order_success', order_id=order.id)
    else:
        form = OrderForm(initial={'status': 'Новый'})

    return render(request, 'orders/create_order.html', {
        'form': form,
        'client': client,
    })


@login_required
@only_operators
def create_client_view(request):
    if request.method == 'POST':
        form = ClientCreateForm(request.POST)
        if form.is_valid():
            phone = form.cleaned_data['phone']
            existing = CustomUser.objects.filter(phone=phone, role='CLIENT').first()
            if existing:
                return redirect('edit_client', client_id=existing.id)
            client = form.save()
            return redirect('create_order_with_client', client_id=client.id)
    else:
        form = ClientCreateForm()
    return render(request, 'orders/create_client.html', {'form': form})



@login_required
@only_operators
def create_order(request):
    if request.method == 'POST':
        form = OrderForm(request.POST)
        if form.is_valid():
            phone = form.cleaned_data['client_phone']
            try:
                client = CustomUser.objects.get(phone=phone, role='CLIENT')
            except CustomUser.DoesNotExist:
                form.add_error('client_phone', 'Клиент с таким телефоном не найден.')
            else:
                order = form.save(commit=False)
                order.status = 'Новый'
                order.client = client
                order.save()

                # Обновим адрес клиента, если он изменился
                new_address = form.cleaned_data.get("client_address")
                if new_address and new_address != client.address:
                    client.address = new_address
                    client.save()

                return redirect('order_success', order_id=order.id)
    else:
        form = OrderForm(initial={'status': 'Новый'})

    return render(request, 'orders/create_order.html', {'form': form})


class OrderEditForm(forms.ModelForm):
    class Meta:
        model = Order
        exclude = ['client']
        labels = {
            'status': 'Статус',
            'total_price': 'Общая стоимость',
        }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for field in self.fields.values():
            field.widget.attrs.update({
                'class': 'w-full border rounded px-3 py-2 bg-gray-50 focus:outline-none focus:ring-2 focus:ring-blue-500'
            })
        self.fields['return_date'].widget.attrs['readonly'] = True
        self.fields['return_date'].widget.attrs['disabled'] = True


@login_required
@only_operators
def edit_order(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    RugFormSet = formset_factory(RugForm, extra=0, can_delete=True)
    existing_rugs = Rug.objects.filter(order=order)
    empty_rug_form = None

    if request.method == 'POST':
        old_status = order.status
        form = OrderEditForm(request.POST, instance=order)
        rug_forms = RugFormSet(request.POST)

        if form.is_valid() and rug_forms.is_valid():
            form.save()
            order.rugs.all().delete()

            valid_forms = [
                f for f in rug_forms.forms
                if not f.cleaned_data.get('DELETE') and (
                        f.cleaned_data.get('rug_type') or f.cleaned_data.get('width') or f.cleaned_data.get('length')
                )
            ]

            order.rug_count = len(valid_forms)
            if order.status == "Доставлен клиенту":
                order.return_date = localdate()

            order.save()

            for rug_form in valid_forms:
                rug = rug_form.save(commit=False)
                rug.order = order
                rug.save()

            if order.status != old_status:
                now_local = localtime(now())

                already_exists = Notification.objects.filter(
                    user=order.client,
                    order=order,
                    message=f"Статус вашего заказа #{order.id} изменён на «{order.status}»"
                ).order_by('-created_at').first()

                if not already_exists or (now_local - already_exists.created_at).total_seconds() > 180:
                    Notification.objects.create(
                        user=order.client,
                        order=order,
                        message=f"Статус вашего заказа #{order.id} изменён на «{order.status}»"
                    )

            return redirect('operator_order_list')
    else:
        initial_data = list(existing_rugs.values('width', 'length'))

        rug_forms = RugFormSet(initial=initial_data)
        empty_rug_form = rug_forms.empty_form
        form = OrderEditForm(instance=order)

    return render(request, 'orders/edit_order.html', {
        'form': form,
        'rug_forms': rug_forms,
        'empty_form': empty_rug_form,
        'order': order
    })


@login_required
def client_search_view(request):
    query = request.GET.get('q', '')
    results = []

    if query:
        clients = CustomUser.objects.filter(
            role='CLIENT',
            phone__icontains=query
        )[:10]

        results = [{
            'id': client.id,
            'label': f"{client.phone} — {client.full_name}",
            'phone': client.phone
        } for client in clients]

    return JsonResponse(results, safe=False)


@login_required
def get_client_address(request):
    phone = request.GET.get("phone")
    try:
        client = CustomUser.objects.get(phone=phone, role="CLIENT")
        return JsonResponse({"address": client.address})
    except CustomUser.DoesNotExist:
        return JsonResponse({"address": ""})


def short_dashboard_redirect(request, short_token):
    token = get_object_or_404(AccessToken, token__startswith=short_token)
    return redirect(f'/dashboard/{token.token}/')


from django.shortcuts import redirect, get_object_or_404
from .models import AccessToken


def redirect_by_token(request):
    short_token = request.GET.get("token")
    token_obj = get_object_or_404(AccessToken, token__startswith=short_token)
    return redirect("client_dashboard", token=token_obj.token)


from django.core.exceptions import PermissionDenied
from django.utils import timezone
from django.views.decorators.http import require_POST


def courier_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or request.user.role != 'COURIER':
            raise PermissionDenied
        return view_func(request, *args, **kwargs)

    return wrapper


@login_required
@courier_required
def courier_dashboard(request):
    orders = Order.objects.filter(courier=request.user).order_by('-created_at')
    return render(request, 'courier/dashboard.html', {'orders': orders})


@require_POST
@login_required
@courier_required
def mark_picked_up(request, order_id):
    order = get_object_or_404(Order, id=order_id, courier=request.user, status='Новый')
    order.status = 'Оценка'
    order.save()
    return redirect('courier_dashboard')


@require_POST
@login_required
@courier_required
def mark_delivered(request, order_id):
    order = get_object_or_404(Order, id=order_id, courier=request.user, status='Готов к возврату')
    order.status = 'Доставлен клиенту'
    order.return_date = timezone.localdate()
    order.save()
    return redirect('courier_dashboard')


# views.py
from django.http import JsonResponse
from .models import Order, CustomUser
from django.views.decorators.http import require_GET
from django.utils.dateparse import parse_date

@require_GET
def get_courier_load(request):
    date_str = request.GET.get("date")
    if not date_str:
        return JsonResponse({"error": "date is required"}, status=400)

    date = parse_date(date_str)
    if not date:
        return JsonResponse({"error": "invalid date"}, status=400)

    couriers = CustomUser.objects.filter(role='COURIER')
    data = {}

    for courier in couriers:
        count = Order.objects.filter(courier=courier, delivery_date=date).count()
        data[courier.id] = {
            "full_name": courier.full_name or courier.username,
            "load": count
        }

    return JsonResponse(data)


from accounts.models import CustomUser

@login_required
@only_operators
def operator_employee_list(request):
    employees = CustomUser.objects.filter(role__in=['OPERATOR', 'COURIER'])

    return render(request, 'orders/employee_list.html', {'employees': employees})


from accounts.models import CustomUser
from .models import ShiftLog
from django.utils import timezone

@login_required
@only_operators
def operator_shift_log(request):
    date_str = request.GET.get('date') or request.POST.get('date')
    selected_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else timezone.now().date()

    employees = CustomUser.objects.exclude(role='CLIENT')

    if request.method == 'POST':
        checked_ids = request.POST.getlist('worked')
        ShiftLog.objects.filter(date=selected_date).exclude(user__id__in=checked_ids).delete()
        for emp_id in checked_ids:
            ShiftLog.objects.get_or_create(user_id=emp_id, date=selected_date)
        return redirect(f'{request.path}?date={selected_date}')

    shiftlogs = ShiftLog.objects.filter(date=selected_date).values_list('user_id', flat=True)
    return render(request, 'orders/shift_log.html', {
        'employees': employees,
        'shiftlogs': shiftlogs,
        'selected_date': selected_date,
    })


@login_required
@only_operators
def operator_shift_report(request):
    from accounts.models import CustomUser
    from orders.models import ShiftLog
    import calendar

    employees = CustomUser.objects.exclude(role='CLIENT')
    employee_id = request.GET.get('employee')
    month_str = request.GET.get('month', timezone.now().strftime('%Y-%m'))

    selected_month = datetime.strptime(month_str, '%Y-%m').date()
    selected_employee = None
    shift_days = []

    if employee_id:
        selected_employee = CustomUser.objects.get(id=employee_id)
        start_date = selected_month.replace(day=1)
        end_day = monthrange(start_date.year, start_date.month)[1]
        end_date = selected_month.replace(day=end_day)

        logs = ShiftLog.objects.filter(user=employee_id, date__range=(start_date, end_date)).values_list('date', flat=True)
        shift_days = [d.day for d in logs]

    # Формируем список дней месяца
    _, last_day = monthrange(selected_month.year, selected_month.month)
    calendar_days = [
        {'date': date(selected_month.year, selected_month.month, day), 'working': day in shift_days}
        for day in range(1, last_day + 1)
    ]

    selected_month_verbose = f"{calendar.month_name[selected_month.month]} {selected_month.year}"

    context = {
        'employees': employees,
        'selected_employee': selected_employee,
        'selected_employee_id': int(employee_id) if employee_id else None,
        'selected_month': month_str,
        'selected_month_verbose': selected_month_verbose,
        'calendar_days': calendar_days,
    }
    if request.GET.get("export") == "1" and selected_employee:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Смены"
        ws.append(["Дата", "Сотрудник", "Роль", "Присутствие"])
        for day in calendar_days:
            ws.append([
                day["date"].strftime("%Y-%m-%d"),
                selected_employee.full_name or selected_employee.username,
                selected_employee.get_role_display(),
                "Да" if day["working"] else "Нет"
            ])
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename=shift_report_{selected_month.strftime("%Y_%m")}.xlsx'
        wb.save(response)
        return response

    return render(request, 'orders/shift_report.html', context)

from django.http import HttpResponse
from django.utils.dateparse import parse_date
from openpyxl import Workbook
from .models import Order
from django.contrib.auth.decorators import login_required
from accounts.models import CustomUser
from django.shortcuts import render

@login_required
def export_orders_excel(request):
    start_str = request.GET.get("start")
    end_str = request.GET.get("end")
    if not start_str or not end_str:
        return render(request, "orders/export_orders.html")

    start = parse_date(start_str)
    end = parse_date(end_str)
    if not start or not end:
        return render(request, "orders/export_orders.html")

    orders = Order.objects.filter(created_at__date__range=(start, end)).select_related("client")
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(["ID", "Клиент", "Телефон", "Адрес", "Дата", "Статус", "Стоимость"])
    for order in orders:
        ws.append([
            order.id,
            order.client.full_name or order.client.username,
            order.client.phone,
            order.client.address,
            order.created_at.strftime("%Y-%m-%d %H:%M"),
            order.status,
            order.total_price
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="orders.xlsx"'
    wb.save(response)
    return response

@login_required
def export_orders_page(request):
    return render(request, "orders/export_orders.html")


# views.py (дополнительно)
from django.contrib.auth.decorators import login_required
from accounts.models import CustomUser
from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import Q
from django import forms
from .models import Order
from openpyxl import Workbook
from django.http import HttpResponse
from django.utils.dateparse import parse_date

class ClientEditForm(forms.ModelForm):
    class Meta:
        model = CustomUser
        fields = ['full_name', 'phone', 'email', 'address']

@login_required
@only_operators
def client_list_view(request):
    query = request.GET.get("q", "")
    clients = CustomUser.objects.filter(role="CLIENT")
    if query:
        clients = clients.filter(
            Q(full_name__icontains=query) | Q(phone__icontains=query)
        )
    return render(request, "orders/client_list.html", {"clients": clients, "q": query})

@login_required
@only_operators
def client_edit_view(request, client_id):
    client = get_object_or_404(CustomUser, id=client_id, role="CLIENT")
    if request.method == "POST":
        form = ClientEditForm(request.POST, instance=client)
        if form.is_valid():
            form.save()
            return redirect("client_list")
    else:
        form = ClientEditForm(instance=client)
    return render(request, "orders/client_edit.html", {"form": form, "client": client})

@login_required
@only_operators
def client_report_view(request):
    client_id = request.GET.get("client")
    start = parse_date(request.GET.get("start") or "1900-01-01")
    end = parse_date(request.GET.get("end") or "2100-01-01")
    client = get_object_or_404(CustomUser, id=client_id, role="CLIENT")
    orders = Order.objects.filter(client=client, created_at__date__range=(start, end))

    if request.GET.get("export") == "1":
        wb = Workbook()
        ws = wb.active
        ws.title = "Client Orders"
        ws.append(["ID", "Дата", "Статус", "Стоимость"])
        for order in orders:
            ws.append([order.id, order.created_at.strftime("%Y-%m-%d"), order.status, order.total_price])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="client_{client.id}_orders.xlsx"'
        wb.save(response)
        return response

    return render(request, "orders/client_report.html", {
        "client": client,
        "orders": orders,
        "start": start,
        "end": end
    })


from django.shortcuts import render, get_object_or_404, redirect
from accounts.models import CustomUser
from .forms import ClientCreateForm
from django.contrib.auth.decorators import login_required


@login_required
@only_operators
def edit_client(request, client_id):
    client = get_object_or_404(CustomUser, id=client_id, role='CLIENT')
    if request.method == 'POST':
        form = ClientCreateForm(request.POST, instance=client)
        if form.is_valid():
            form.save()
            return redirect('client_list')
    else:
        form = ClientCreateForm(instance=client)

    return render(request, 'orders/edit_client.html', {
        'form': form,
        'client': client
    })


from django.db.models import Count, Sum

@login_required
@only_operators
def client_report_view(request, client_id):
    client = get_object_or_404(CustomUser, id=client_id, role='CLIENT')
    orders = Order.objects.filter(client=client).order_by('-created_at')

    total_orders = orders.count()
    total_paid = orders.aggregate(Sum('total_price'))['total_price__sum'] or 0

    return render(request, 'orders/client_report.html', {
        'client': client,
        'orders': orders,
        'total_orders': total_orders,
        'total_paid': total_paid,
    })
